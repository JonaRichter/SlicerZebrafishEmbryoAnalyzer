import os
import pytest


RESULTS = [
    {
        "filename": "fish1.png", "length": 1200.0, "curvature": 2,
        "ratio": 1.05, "eye_area": None, "eye_diameter": None, "error": None,
    },
    {
        "filename": "fish2.png", "length": None, "curvature": None,
        "ratio": None, "eye_area": None, "eye_diameter": None, "error": "Segmentation failed",
    },
]


def test_export_excel_creates_file(tmp_path):
    from ZebrafishEmbryoAnalyzerLib.export import export_excel
    out = str(tmp_path / "results.xlsx")
    export_excel(RESULTS, out)
    assert os.path.exists(out)
    assert os.path.getsize(out) > 0


def test_export_csv_creates_file(tmp_path):
    from ZebrafishEmbryoAnalyzerLib.export import export_csv
    out = str(tmp_path / "results.csv")
    export_csv(RESULTS, out)
    assert os.path.exists(out)
    lines = open(out).readlines()
    assert len(lines) == 3  # header + 2 rows


def test_export_excel_includes_error_column(tmp_path):
    import openpyxl
    from ZebrafishEmbryoAnalyzerLib.export import export_excel
    out = str(tmp_path / "results.xlsx")
    export_excel(RESULTS, out)
    wb = openpyxl.load_workbook(out)
    headers = [cell.value for cell in wb.active[1]]
    assert "Error" in headers


def test_export_filtered_skips_excluded(tmp_path):
    from ZebrafishEmbryoAnalyzerLib.export import export_excel
    import openpyxl
    out = str(tmp_path / "filtered.xlsx")
    filtered = [r for r in RESULTS if r["error"] is None]
    export_excel(filtered, out)
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    # Row 1 = header, row 2 = the one result row. Statistics/class-distribution/
    # boxplot sections (issue #75) are appended after, so max_row now exceeds 2 —
    # what matters is the data rows themselves, not the sheet's total length.
    assert ws.cell(row=2, column=1).value == "fish1.png"
    assert ws.max_row > 2


def test_metric_keys_derived_from_headers_excludes_filename_and_error():
    from ZebrafishEmbryoAnalyzerLib.export import METRIC_KEYS, HEADERS
    keys_from_headers = {k for _, k in HEADERS}
    assert set(METRIC_KEYS) == keys_from_headers - {"filename", "error"}
    assert "filename" not in METRIC_KEYS
    assert "error" not in METRIC_KEYS


def test_export_excel_renders_excluded_cell_not_dropped_row(tmp_path):
    """Issue #74/#75: a per-metric-excluded cell must render as literal
    "Excluded" text, and the row must still be present — not dropped."""
    from ZebrafishEmbryoAnalyzerLib.export import export_excel, HEADERS
    import openpyxl
    out = str(tmp_path / "partial_exclude.xlsx")
    excluded = {"fish1.png": {"curvature"}}
    export_excel(RESULTS, out, excluded=excluded)
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    # Rows 2-3 = the two result rows, in order, none dropped. (Issue #75
    # appends statistics/class-distribution/boxplot sections after these,
    # so max_row now exceeds 3 — that's expected, not a regression.)
    assert ws.cell(row=3, column=1).value == "fish2.png"
    header_row = [c.value for c in ws[1]]
    curvature_col = header_row.index("Curvature class") + 1
    length_col = header_row.index("Length (µm)") + 1
    assert ws.cell(row=2, column=curvature_col).value == "Excluded"
    assert ws.cell(row=2, column=length_col).value == 1200.0  # untouched


def test_export_csv_renders_excluded_cell_not_dropped_row(tmp_path):
    from ZebrafishEmbryoAnalyzerLib.export import export_csv
    out = str(tmp_path / "partial_exclude.csv")
    excluded = {"fish1.png": {"curvature"}}
    export_csv(RESULTS, out, excluded=excluded)
    lines = open(out).readlines()
    assert len(lines) == 3  # header + both rows
    assert "Excluded" in lines[1]


# ---------------------------------------------------------------------------
# Issue #75: Excel export depth — boxplots, stats, class distribution, sheet name
# ---------------------------------------------------------------------------

_STATS_RESULTS = [
    {"filename": f"fish{i}.png", "length": float(100 + i * 10), "curvature": (i % 4) + 1,
     "ratio": 1.0 + i * 0.05, "eye_area": None, "eye_diameter": None,
     "edema_area": None, "swim_area": None, "swim_width": None, "error": None}
    for i in range(5)
]


def test_sanitize_sheet_name_strips_forbidden_chars_and_caps_length():
    from ZebrafishEmbryoAnalyzerLib.export import sanitize_sheet_name
    assert sanitize_sheet_name("A/B\\C?D*E[F]G:H'I") == "ABCDEFGHI"
    long_name = "x" * 50
    assert len(sanitize_sheet_name(long_name)) == 31
    assert sanitize_sheet_name("") == "Zebrafish Results"
    assert sanitize_sheet_name(None) == "Zebrafish Results"
    assert sanitize_sheet_name("  ") == "Zebrafish Results"
    assert sanitize_sheet_name("My Sheet") == "My Sheet"


def test_export_excel_uses_custom_sanitized_sheet_name(tmp_path):
    import openpyxl
    from ZebrafishEmbryoAnalyzerLib.export import export_excel
    out = str(tmp_path / "sheet.xlsx")
    export_excel(RESULTS, out, sheet_name="Batch 2026-07/20")
    wb = openpyxl.load_workbook(out)
    assert wb.active.title == "Batch 2026-0720"  # "/" is stripped, not replaced


def test_export_excel_default_sheet_name_when_none_given(tmp_path):
    import openpyxl
    from ZebrafishEmbryoAnalyzerLib.export import export_excel
    out = str(tmp_path / "sheet_default.xlsx")
    export_excel(RESULTS, out)
    wb = openpyxl.load_workbook(out)
    assert wb.active.title == "Zebrafish Results"


def test_stats_computes_median_percentiles_mean_stdev():
    from ZebrafishEmbryoAnalyzerLib.export import _stats
    med, p25, p75, mean, std = _stats([1.0, 2.0, 3.0, 4.0, 5.0])
    assert med == pytest.approx(3.0)
    assert mean == pytest.approx(3.0)
    assert p25 == pytest.approx(2.0)
    assert p75 == pytest.approx(4.0)
    assert std > 0


def test_stats_empty_returns_na_tuple():
    from ZebrafishEmbryoAnalyzerLib.export import _stats
    assert _stats([]) == ("N/A",) * 5


def test_clean_numeric_values_excludes_per_metric_and_error_rows():
    from ZebrafishEmbryoAnalyzerLib.export import _clean_numeric_values
    results = [
        {"filename": "a.png", "length": 100.0, "error": None},
        {"filename": "b.png", "length": 200.0, "error": None},
        {"filename": "c.png", "length": 300.0, "error": "boom"},
    ]
    excluded = {"a.png": {"length"}}
    values = _clean_numeric_values(results, excluded, "length")
    assert values == [200.0]  # a excluded, c errored, only b counts


def test_export_excel_statistics_section_ignores_excluded_values(tmp_path):
    """Issue #75's core acceptance criterion: statistics must be
    exclusion-aware, using the per-metric exclude model from issue #74."""
    import openpyxl
    from ZebrafishEmbryoAnalyzerLib.export import export_excel
    out = str(tmp_path / "stats.xlsx")
    # Exclude the lowest length value; the median without it should shift up.
    excluded = {"fish0.png": {"length"}}
    export_excel(_STATS_RESULTS, out, excluded=excluded)
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    rows = [[c.value for c in row] for row in ws.iter_rows()]
    median_row = next(r for r in rows if r and r[0] == "Median Length (µm)")
    # fish0 (length=100) excluded -> remaining [110,120,130,140], median=125
    assert median_row[1] == pytest.approx(125.0)


def test_export_excel_class_distribution_section_present(tmp_path):
    import openpyxl
    from ZebrafishEmbryoAnalyzerLib.export import export_excel
    out = str(tmp_path / "classdist.xlsx")
    export_excel(_STATS_RESULTS, out)
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    rows = [[c.value for c in row] for row in ws.iter_rows()]
    assert any(r and r[0] == "Class Distribution" for r in rows)
    assert any(r and r[0] == "Class 1" for r in rows)


def test_export_excel_embeds_boxplot_image(tmp_path):
    import openpyxl
    from ZebrafishEmbryoAnalyzerLib.export import export_excel
    out = str(tmp_path / "boxplot.xlsx")
    export_excel(_STATS_RESULTS, out)
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    assert len(ws._images) >= 1


def test_make_boxplots_image_returns_nonempty_png_bytes():
    from ZebrafishEmbryoAnalyzerLib.export import _make_boxplots_image
    png_bytes = _make_boxplots_image(_STATS_RESULTS, {})
    assert isinstance(png_bytes, bytes)
    assert len(png_bytes) > 0
    assert png_bytes[:8] == b"\x89PNG\r\n\x1a\n"  # PNG file signature


def test_class_distribution_counts_classes_and_uncertain():
    from ZebrafishEmbryoAnalyzerLib.export import _class_distribution
    results = [
        {"filename": "a.png", "curvature": 1, "error": None},
        {"filename": "b.png", "curvature": 1, "error": None},
        {"filename": "c.png", "curvature": 5, "error": None},  # uncertain/not classified
        {"filename": "d.png", "curvature": None, "error": None},  # skipped
        {"filename": "e.png", "curvature": 2, "error": "boom"},  # errored, skipped
    ]
    dist = dict(_class_distribution(results, {}))
    assert dist["Class 1"] == 2
    assert dist["Not Classified"] == 1
    assert dist["Class 2"] == 0
